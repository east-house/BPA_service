from fastapi import HTTPException, File, status
from openpyxl import load_workbook
from loguru import logger
import json
import re


def check_save_xlsx(target_file: File) -> str:
    contents = check_xlsx(target_file)
    save_path = save_xlsx(target_file, contents)
    return save_path


def check_xlsx(check_file: File) -> None:
    if not check_file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="The file extension is invalid.")

    from io import BytesIO
    import pandas as pd

    contents = check_file.file.read()
    try:
        WB = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="The file extension is invalid.")

    try:
        SHEETLEN = len(WB.sheetnames)
        if SHEETLEN > 1:
            logger.warning("시트가 2개이상입니다. 첫번째 시트를 대상으로 양식 확인합니다.")
            WS = WB.active
        else:
            WS = WB.active

        # 데이터프레임 지정 및 전처리
        df = pd.DataFrame(WS.values)
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        cleaned_df = df.reset_index(drop=True)

        # 행을 하나씩 읽어서, NaN값이 하나도 없는 행을 찾음.
        column_row_index = None
        for idx, row in cleaned_df.iterrows():
            nan_count = row.isna().sum()
            if nan_count == 0:
                column_row_index = idx
                break
            else:
                pass

        # 실제 표(데이터가 담긴 영역)를 찾아서 데이터프레임으로 변환
        if column_row_index is not None:
            DF_FINAL = cleaned_df.iloc[column_row_index:].reset_index(
                drop=True)
            DF_FINAL.columns = DF_FINAL.iloc[0]
            DF_FINAL = DF_FINAL[1:].reset_index(drop=True)
        else:
            DF_FINAL = pd.DataFrame()

        row_shape, column_shape = DF_FINAL.shape

        if (row_shape >= 2) and (column_shape >= 1):
            logger.warning(f"check xlsx done!!")
            return contents
        else:
            HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                          detail="The Excel format is incorrect. Please check the format of the uploaded Excel file.")
    except:
        HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                      detail="The Excel format is incorrect. Please check the format of the uploaded Excel file.")


def save_xlsx(save_file: File, contents: str) -> str:
    save_file_path = f"imports/{save_file.filename}"
    # contents = save_file.file.read()
    logger.warning(f"save_file_path : {save_file_path}")
    with open(save_file_path, "wb") as buffer:
        buffer.write(contents)
    return save_file_path


def check_yes(text):
    if re.search(r'yes', text, re.IGNORECASE):
        return "YES"
    else:
        return "NO"


def extract_python_code(text):
    python_code_pattern = re.compile(r'```python\n(.*?)\n```', re.DOTALL)
    match = python_code_pattern.search(text)
    if match:
        return match.group(1).strip()
    else:
        return ""


def extract_python_pyplot_code(text):
    python_code_pattern = re.compile(
        r'```python\n(.*?)\nplt.show()', re.DOTALL)
    match = python_code_pattern.search(text)
    if match:
        return match.group(1).strip()
    else:
        return ""


def extract_python_plotly_code(text):
    python_code_pattern = re.compile(
        r'```python\n(.*?)\nfig.show()', re.DOTALL)
    match = python_code_pattern.search(text)
    if match:
        return match.group(1).strip()
    else:
        return ""


def format_number(x):
    '''
        Des:
            숫자형 데이터 comma로 구분
    '''
    if isinstance(x, (int, float)):  # 숫자형 데이터인지 확인
        return "{:,}".format(x)
    else:
        return x


def wrapped_event(message_id='', session_id='', type='', source=[], data='', plot='',  relevant_query=[], download=''):
    return json.dumps({
        "state": status.HTTP_200_OK,
        "result": {
            "message_id": message_id,
            "session_id": session_id,
            "type": type,
            "source": [{
                "href": "http://www.shippingnewsnet.com/news/articleView.html?idxno=45703",
                "title": "Bpa, '22년 부산항 컨테이너 물동량 2,350만teu 목표 < 항만 < 뉴스 < 기사본문 - 쉬핑뉴스넷"
            },
                {
                "href": "https://www.data.go.kr/data/15055477/fileData.do",
                "title": "부산항만공사_부산항 연도별 물동량 추이_20231231 | 공공데이터포털"
            },
                {
                "href": "https://www.data.go.kr/data/15055478/fileData.do",
                "title": "부산항만공사_부산항 컨테이너 수송통계_20221231 | 공공데이터포털"
            },
                {
                "href": "https://www.yna.co.kr/view/AKR20221025051900051",
                "title": "심상찮은 부산항 물동량 감소세…9월에만 14% ↓ | 연합뉴스"
            }],
            "data": data,
            "plot": plot,
            "relevant_query": relevant_query,
            "download": download
        },
        "message": "success"
    }, ensure_ascii=False)


def wrapped_string_generator(string_response, session_id, type, message_id: str = 1):
    if isinstance(string_response, list):
        string_response_form = string_response
    else:
        string_response_form = [string_response]
    for string in string_response_form:
        yield json.dumps({
            "state": status.HTTP_200_OK,
            "result": {
                "message_id": message_id,
                "session_id": session_id,
                "type": type,
                "source": [],
                "data": string,
                "plot": "",
                "relevant_query": [],
                "download": ""
            },
            "message": "success"
        }, ensure_ascii=False)
