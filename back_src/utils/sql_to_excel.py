import re
from . import *
from typing import Union, List
from modules.chain import Chain__summary
from utils.util import format_number


class ExcelTemplateApplier:
    def __init__(self, MESSAGE):
        self.MESSAGE: str = MESSAGE
        self.DF_FINAL: pd.DataFrame = pd.DataFrame()
        self.COLS: List[str] = []
        self.SAMPLES: List[str] = []
        self.LLM_COLS: List[str] = []
        self.INFO_DEFAULT: List[str] = []
        self.TABLE_INFO: List[str] = []

    def apply_format(self, DF: pd.DataFrame):
        '''
            Des:
                숫자형 데이터 comma로 구분
                - 연도 컬럼은 포맷 적용 생략
            Args:
                적용할 데이터 프레임
        '''
        for column in DF.columns:
            if column not in ["년도", "연", "연도", "년"]:
                DF[column] = DF[column].apply(format_number)
        self.DF = DF

    def set_data(self):
        '''
            Des:
                데이터 지정
                - SAMPLES : 값 리스트
                - ONE_SAMPLES : 값 한개 (양식 미업로드시 컬럼 너비 계산에 사용)
                - LLM_COLS : 컬럼 리스트
        '''
        self.SAMPLES = self.DF.values.tolist()

        # 데이터가 2개이상일 경우 self.SAMPLES는 이중리스트
        if type(self.SAMPLES[0]) == list:
            self.ONE_SAMPLE = self.SAMPLES[0]
        else:
            self.ONE_SAMPLE = copy.deepcopy(self.SAMPLES)

        self.LLM_COLS = self.DF.columns.tolist()

    def set_time_and_path(self, NOW):
        '''
            Des:
                저장 경로 지정
            Args:
                현재 시간
        '''
        formatted_date = NOW.strftime('%Y%m%d')
        self.SAVE_PATH = f'exports/SQL2Excel/{formatted_date}_통계추출결과'

    def extract_common_parts(self):
        '''
            Des:
                공통 내용 지정
                - INFO_DEFAULT : 내용 요약 & 포트미스 자료와 상이할 수 있음
        '''
        INFO_MESSAGE = Chain__summary.invoke({"MESSAGE": self.MESSAGE}).content
        self.INFO_DEFAULT = [f" * {INFO_MESSAGE}",
                             " * 통합포트미스의 자료와 상이할 수 있습니다."]

    def yes_template_common_parts(self, PATH):
        '''
            Des:
                양식 업로드시, 기본 정보 추출
                - WB (Workbook)
                - WS (WorkSheet)
                - DF_FINAL : 실제 데이터가 담긴 영역
                - COLS : 실제 데이터 영역의 컬럼
                - TABLE_INFO : 실제 데이터 영역 정보 (행번호, 컬럼 시작번호, 컬럼 끝 번호)
            Args:
                PATH : Chainlit 내부 업로드된 파일 경로
        '''

        self.WB = openpyxl.load_workbook(PATH, data_only=True)
        SHEETLEN = len(self.WB.sheetnames)
        if SHEETLEN > 1:
            logger.warning("시트가 2개이상입니다. 첫번째 시트를 대상으로 양식 확인합니다.")
            self.WS = self.WB.active
        else:
            self.WS = self.WB.active

        # 데이터프레임 지정 및 전처리
        df = pd.DataFrame(self.WS.values)
        df = df.dropna(how='all')
        df = df.dropna(axis=1, how='all')
        cleaned_df = df.reset_index(drop=True)

        # 행을 하나씩 읽어서, NaN값이 하나도 없는 행을 찾음.
        column_row_index = None
        for idx, row in cleaned_df.iterrows():
            nan_count = row.isna().sum()
            if nan_count == 0:
                column_row_index = idx
                break
            else:
                pass

        # 실제 표(데이터가 담긴 영역)를 찾아서 데이터프레임으로 변환
        if column_row_index is not None:
            DF_FINAL = cleaned_df.iloc[column_row_index:].reset_index(
                drop=True)
            DF_FINAL.columns = DF_FINAL.iloc[0]
            self.DF_FINAL = DF_FINAL[1:].reset_index(drop=True)
        else:
            self.DF_FINAL = pd.DataFrame()

        # 실제 표의 컬럼, 행위치를 저장
        self.COLS = list(self.DF_FINAL.columns)
        for row_idx, row in enumerate(self.WS.iter_rows(values_only=True)):
            a = set(list(row))
            b = set(list(self.COLS))
            is_subset = b.issubset(a)
            if is_subset:
                start_col_idx = list(~pd.array(row).isna()).index(True)
                end_col_idx = len(row)-1
                self.TABLE_INFO = [row_idx, start_col_idx, end_col_idx]
                break

    def no_template(self):

        WB = Workbook()
        WS = WB.active

        # 컬럼 삽입 & 스타일
        for num, val in enumerate(self.LLM_COLS):
            alpha = get_column_letter(num+2)
            WS[f"{alpha}7"] = val
            WS[f"{alpha}7"].fill = PatternFill(
                start_color="cae8ff", end_color="cae8ff", fill_type="solid")
            WS[f"{alpha}7"].font = Font(bold=True)
            WS[f"{alpha}7"].alignment = Alignment(
                horizontal="center", vertical="center")
            WS.column_dimensions[alpha].width = 3 + \
                select_width(val, str(self.ONE_SAMPLE[num]))

        # 데이터삽입
        START_IDX = 8
        for row in tqdm(self.SAMPLES):
            for num, val in enumerate(row):
                alpha = get_column_letter(num+2)
                WS[f"{alpha}{START_IDX}"] = val
            START_IDX += 1

        # 요약문 추가
        for idx, info in enumerate(self.INFO_DEFAULT):
            WS[f"B{idx+2}"] = info
            if idx+1 == len(self.INFO_DEFAULT):
                WS[f"B{idx+2}"].font = Font(color="FF0000", bold=True)

        NEW_SAVE_PATH = f"{self.SAVE_PATH}_양식없음.xlsx"
        WB.save(NEW_SAVE_PATH)
        return NEW_SAVE_PATH

    def yes_template_with_template_cols(self):
        '''
            Des:
                양식 컬럼 기반 엑셀 생성
        '''

        # 표 컬럼, 행 위치 저장
        TABLE_COL_START_ALPHA = get_column_letter(self.TABLE_INFO[1]+1)
        TABLE_COL_START_IDX = (self.TABLE_INFO[1]+1)
        TABLE_COL_END_ALPHA = get_column_letter(self.TABLE_INFO[2]+1)
        TABLE_COL_END_IDX = (self.TABLE_INFO[2]+1)
        TABLE_ROW_START_IDX = (self.TABLE_INFO[0]+1)
        # print(f"컬럼 시작 알파벳 : {TABLE_COL_START_ALPHA} <==> 컬럼 시작 인덱스 : {TABLE_COL_START_IDX}")
        # print(f"컬럼 종료 알파벳 : {TABLE_COL_END_ALPHA} <==> 컬럼 종료 인덱스 : {TABLE_COL_END_IDX}")
        # print(f"행 인덱스 : {TABLE_ROW_START_IDX}")

        # 컬럼 알파벳 목록 추출
        COLS_ALPHA_ALL = get_alphabets_between(
            TABLE_COL_START_ALPHA, TABLE_COL_END_ALPHA)

        # 모든 컬럼의 너비정보가 있도록 설정 (일부 컬럼의 너비정보가 빠진 경우가 있음)
        for alpha in COLS_ALPHA_ALL:
            if alpha in list(self.WS.column_dimensions.keys()):
                pass
            else:
                self.WS.column_dimensions[alpha] = self.WS.column_dimensions[chr(
                    ord(alpha)-1)]

        # 셀 위치별 스타일 저장
        USE_CELLS = [f'{alpha}{TABLE_ROW_START_IDX}' for alpha in COLS_ALPHA_ALL] + \
                    [f'{alpha}{TABLE_ROW_START_IDX+1}' for alpha in COLS_ALPHA_ALL]
        CELL_STYLES = {}
        for row in self.WS.iter_rows():
            for cell in row:
                if cell.coordinate in USE_CELLS:
                    CELL_STYLES[cell.coordinate] = {
                        "font": copy.copy(cell.font),
                        "fill": copy.copy(cell.fill),
                        "border": copy.copy(cell.border),
                        "alignment": copy.copy(cell.alignment),
                        "width": self.WS.column_dimensions[re.sub(r'\d', '', cell.coordinate)].width
                    }
                if cell.coordinate == f"{TABLE_COL_END_ALPHA}{int(TABLE_ROW_START_IDX)+1}":
                    break
            if cell.coordinate == f"{TABLE_COL_END_ALPHA}{int(TABLE_ROW_START_IDX)+1}":
                break

        # 고정된 위치(B7 기준)로 이동
        FIXED_POS = "B7"
        BEFORE_POS = list(CELL_STYLES.keys())[0]

        COL_DIFF, ROW_DIFF = calculate_difference(BEFORE_POS, FIXED_POS)
        NEW_TABLE_COL_START_IDX = TABLE_COL_START_IDX+COL_DIFF
        NEW_TABLE_COL_END_IDX = TABLE_COL_END_IDX+COL_DIFF
        NEW_TABLE_ROW_START_IDX = TABLE_ROW_START_IDX+ROW_DIFF

        NEW_CELL_STYLES = {}
        for k, v in CELL_STYLES.items():
            NEW_POS = f"{chr(ord(split_cell(k)[0])+COL_DIFF)}{split_cell(k)[1]+ROW_DIFF}"
            NEW_CELL_STYLES[NEW_POS] = v

        # 양식 기본 적용
        from openpyxl import Workbook
        NEW_WB = Workbook()
        NEW_WS = NEW_WB.active

        for col_idx, col_val in zip(range(NEW_TABLE_COL_START_IDX, NEW_TABLE_COL_END_IDX+1), self.COLS):
            alpha = get_column_letter(col_idx)
            AlphaNum = f"{alpha}{NEW_TABLE_ROW_START_IDX}"
            NEW_WS[AlphaNum] = col_val
            NEW_WS[AlphaNum].font = NEW_CELL_STYLES[AlphaNum]["font"]
            NEW_WS[AlphaNum].fill = NEW_CELL_STYLES[AlphaNum]["fill"]
            NEW_WS[AlphaNum].border = NEW_CELL_STYLES[AlphaNum]["border"]
            NEW_WS[AlphaNum].alignment = NEW_CELL_STYLES[AlphaNum]["alignment"]
            NEW_WS.column_dimensions[alpha].width = NEW_CELL_STYLES[AlphaNum]["width"]

        # 값 채우기
        for idx, sample in enumerate(tqdm(self.SAMPLES)):
            for (col_idx, col_val) in zip(range(NEW_TABLE_COL_START_IDX, NEW_TABLE_COL_END_IDX+1), sample):
                alpha = get_column_letter(col_idx)
                VALUE_STYLE = f"{alpha}{NEW_TABLE_ROW_START_IDX+1}"
                AlphaNum = f"{alpha}{NEW_TABLE_ROW_START_IDX+idx+1}"
                NEW_WS[AlphaNum] = col_val
                NEW_WS[AlphaNum].font = NEW_CELL_STYLES[VALUE_STYLE]["font"]
                NEW_WS[AlphaNum].fill = NEW_CELL_STYLES[VALUE_STYLE]["fill"]
                NEW_WS[AlphaNum].border = NEW_CELL_STYLES[VALUE_STYLE]["border"]
                NEW_WS[AlphaNum].alignment = NEW_CELL_STYLES[VALUE_STYLE]["alignment"]

        # 기본값 채우기
        for idx, info in enumerate(self.INFO_DEFAULT):
            NEW_WS[f"B{idx+2}"] = info
            if idx+1 == len(self.INFO_DEFAULT):
                NEW_WS[f"B{idx+2}"].font = Font(color="FF0000", bold=True)

        NEW_SAVE_PATH = f"{self.SAVE_PATH}_양식컬럼사용.xlsx"
        NEW_WB.save(NEW_SAVE_PATH)
        return NEW_SAVE_PATH

    def yes_template_with_llm_cols(self):
        '''
            Des:
                LLM이 생성한 컬럼 기반 엑셀 생성
        '''
        # 표 컬럼, 행 위치 저장
        TABLE_COL_START_ALPHA = get_column_letter(self.TABLE_INFO[1]+1)
        TABLE_COL_START_IDX = (self.TABLE_INFO[1]+1)
        TABLE_COL_END_ALPHA = get_column_letter(self.TABLE_INFO[2]+1)
        TABLE_COL_END_IDX = (self.TABLE_INFO[2]+1)
        TABLE_ROW_START_IDX = (self.TABLE_INFO[0]+1)
        # print(f"컬럼 시작 알파벳 : {TABLE_COL_START_ALPHA} <==> 컬럼 시작 인덱스 : {TABLE_COL_START_IDX}")
        # print(f"컬럼 종료 알파벳 : {TABLE_COL_END_ALPHA} <==> 컬럼 종료 인덱스 : {TABLE_COL_END_IDX}")
        # print(f"행 인덱스 : {TABLE_ROW_START_IDX}")

        # 컬럼 알파벳 목록 추출
        COLS_ALPHA_ALL = get_alphabets_between(
            TABLE_COL_START_ALPHA, TABLE_COL_END_ALPHA)

        # 모든 컬럼의 너비정보가 있도록 설정 (일부 컬럼의 너비정보가 빠진 경우가 있음)
        for alpha in COLS_ALPHA_ALL:
            if alpha in list(self.WS.column_dimensions.keys()):
                pass
            else:
                self.WS.column_dimensions[alpha] = self.WS.column_dimensions[chr(
                    ord(alpha)-1)]

        # 셀 위치별 스타일 저장
        USE_CELLS = [f'{alpha}{TABLE_ROW_START_IDX}' for alpha in COLS_ALPHA_ALL] + \
                    [f'{alpha}{TABLE_ROW_START_IDX+1}' for alpha in COLS_ALPHA_ALL]
        CELL_STYLES = {}
        for row in self.WS.iter_rows():
            for cell in row:
                if cell.coordinate in USE_CELLS:
                    CELL_STYLES[cell.coordinate] = {
                        "font": copy.copy(cell.font),
                        "fill": copy.copy(cell.fill),
                        "border": copy.copy(cell.border),
                        "alignment": copy.copy(cell.alignment),
                        "width": self.WS.column_dimensions[re.sub(r'\d', '', cell.coordinate)].width
                    }
                if cell.coordinate == f"{TABLE_COL_END_ALPHA}{int(TABLE_ROW_START_IDX)+1}":
                    break
            if cell.coordinate == f"{TABLE_COL_END_ALPHA}{int(TABLE_ROW_START_IDX)+1}":
                break

        # 고정된 위치(B7 기준)로 이동
        FIXED_POS = "B7"
        BEFORE_POS = list(CELL_STYLES.keys())[0]
        COL_DIFF, ROW_DIFF = calculate_difference(BEFORE_POS, FIXED_POS)
        NEW_TABLE_COL_START_IDX = TABLE_COL_START_IDX+COL_DIFF
        NEW_TABLE_COL_END_IDX = TABLE_COL_END_IDX+COL_DIFF
        NEW_TABLE_ROW_START_IDX = TABLE_ROW_START_IDX+ROW_DIFF

        NEW_CELL_STYLES = {}
        for k, v in CELL_STYLES.items():
            NEW_POS = f"{chr(ord(split_cell(k)[0])+COL_DIFF)}{split_cell(k)[1]+ROW_DIFF}"
            NEW_CELL_STYLES[NEW_POS] = v

        NEW_WB = Workbook()
        NEW_WS = NEW_WB.active

        # 컬럼 삽입 & 스타일
        for col_idx, col_val in zip(range(NEW_TABLE_COL_START_IDX, NEW_TABLE_COL_END_IDX+1), self.LLM_COLS):
            alpha = get_column_letter(col_idx)
            AlphaNum = f"{alpha}{NEW_TABLE_ROW_START_IDX}"
            NEW_WS[AlphaNum] = col_val
            NEW_WS[AlphaNum].font = NEW_CELL_STYLES[AlphaNum]["font"]
            NEW_WS[AlphaNum].fill = NEW_CELL_STYLES[AlphaNum]["fill"]
            NEW_WS[AlphaNum].border = NEW_CELL_STYLES[AlphaNum]["border"]
            NEW_WS[AlphaNum].alignment = NEW_CELL_STYLES[AlphaNum]["alignment"]
            NEW_WS.column_dimensions[alpha].width = 3 + \
                NEW_CELL_STYLES[AlphaNum]["width"]

        # 값 채우기
        for idx, sample in enumerate(tqdm(self.SAMPLES)):
            for (col_idx, col_val) in zip(range(NEW_TABLE_COL_START_IDX, NEW_TABLE_COL_END_IDX+1), sample):
                alpha = get_column_letter(col_idx)
                VALUE_STYLE = f"{alpha}{NEW_TABLE_ROW_START_IDX+1}"
                AlphaNum = f"{alpha}{NEW_TABLE_ROW_START_IDX+idx+1}"
                NEW_WS[AlphaNum] = col_val
                NEW_WS[AlphaNum].font = NEW_CELL_STYLES[VALUE_STYLE]["font"]
                NEW_WS[AlphaNum].fill = NEW_CELL_STYLES[VALUE_STYLE]["fill"]
                NEW_WS[AlphaNum].border = NEW_CELL_STYLES[VALUE_STYLE]["border"]
                NEW_WS[AlphaNum].alignment = NEW_CELL_STYLES[VALUE_STYLE]["alignment"]

        # 요약문 추가
        for idx, info in enumerate(self.INFO_DEFAULT):
            NEW_WS[f"B{idx+2}"] = info
            if idx+1 == len(self.INFO_DEFAULT):
                NEW_WS[f"B{idx+2}"].font = Font(color="FF0000", bold=True)

        NEW_SAVE_PATH = f"{self.SAVE_PATH}_양식컬럼미사용.xlsx"
        NEW_WB.save(NEW_SAVE_PATH)
        return NEW_SAVE_PATH


korean_pattern = re.compile('[가-힣]')
number_pattern = re.compile('[0-9]')
english_pattern = re.compile('[a-zA-Z]')
space_pattern = re.compile('[ ]')
special_pattern = re.compile('[^가-힣0-9a-zA-Z]')


def calc_width(sample):
    korean_count = len(korean_pattern.findall(sample))
    number_count = len(number_pattern.findall(sample))
    english_count = len(english_pattern.findall(sample))
    space_count = len(space_pattern.findall(sample))
    special_count = len(special_pattern.findall(sample))
    width = (korean_count*2.2)+(number_count*1.2) + \
        (english_count*1.2)+(space_count*1)+(special_count*1.2)
    return width


def select_width(col, sample):
    col_len = calc_width(col)
    sample_len = calc_width(sample)
    return max(col_len, sample_len)


def col_to_index(col):
    """열 알파벳을 숫자로 변환"""
    index = 0
    for char in col:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


def split_cell(cell):
    """주어진 셀 주소에서 열과 행 분리"""
    for i, char in enumerate(cell):
        if char.isdigit():
            return cell[:i], int(cell[i:])
    return cell, None


def calculate_difference(cell1, cell2):
    """두 셀 주소 사이의 차이값 계산"""
    col1, row1 = split_cell(cell1)
    col2, row2 = split_cell(cell2)
    col_index1 = col_to_index(col1)
    col_index2 = col_to_index(col2)
    col_diff = col_index2 - col_index1
    row_diff = row2 - row1
    return col_diff, row_diff


def get_alphabets_between(start, end):
    start_ascii = ord(start)
    end_ascii = ord(end)
    return [chr(i) for i in range(start_ascii, end_ascii+1)]
