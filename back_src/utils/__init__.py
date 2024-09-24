import re
import numpy as np
import pandas as pd
import ast
import os
import json
import copy
import mariadb
from loguru import logger
from typing import Any
from tqdm import tqdm
import requests
from bs4 import BeautifulSoup
from duckduckgo_search import DDGS
from utils.util import check_yes

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook


from langchain.pydantic_v1 import BaseModel, Field
from langchain.schema import BaseOutputParser
from langchain_core.output_parsers import PydanticOutputParser

from configs.template.text_to_sql import Text2SQLConfig
SYSTEM_PROMPT = Text2SQLConfig.qwen_system
